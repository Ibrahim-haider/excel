
from __future__ import annotations

import io
import math
import re
from calendar import monthrange
from datetime import datetime
from pathlib import Path

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

BASE_DIR = Path(__file__).resolve().parent
DATA_FILE = BASE_DIR / "data" / "Ibrahim.xlsx"

st.set_page_config(
    page_title="Executive Sales Intelligence",
    page_icon="📊",
    layout="wide",
)

st.markdown("""
<style>
.block-container {padding-top: 1rem; padding-bottom: 2rem;}
[data-testid="stMetric"] {
    background: #ffffff;
    border: 1px solid #d7deea;
    border-radius: 16px;
    padding: 14px;
    box-shadow: 0 3px 10px rgba(17,24,39,.06);
}
.topbar {
    background:#223f72;
    color:white;
    padding:14px 18px;
    border-radius:6px;
    font-size:1.45rem;
    font-weight:750;
    margin-bottom:18px;
}
.section-title {
    background:#223f72;
    color:white;
    padding:8px 14px;
    border-radius:6px;
    font-weight:700;
    text-align:center;
    margin:10px 0 14px;
}
.note {
    background:#f8fafc;
    border-left:4px solid #223f72;
    padding:10px 14px;
    border-radius:6px;
    color:#475569;
}
.small {font-size:.82rem; color:#64748b;}
</style>
""", unsafe_allow_html=True)


CANONICAL_ALIASES = {
    "Unit": ["unit", "business unit", "division", "region"],
    "Store": ["store", "store code", "branch code", "outlet"],
    "Zone": ["zone", "region", "area", "territory"],
    "Branch Name": ["branch name", "branch", "store name", "outlet name"],
    "Target": ["target", "mtd target", "sales target", "monthly target"],
    "Sale Value": ["sale value", "sales value", "mtd sales", "actual sales", "sales"],
    "YTD TGT": ["ytd tgt", "ytd target", "year to date target"],
    "YTD SV": ["ytd sv", "ytd sales", "year to date sales"],
    "CS TGT": ["cs tgt", "cash target", "cash sales target"],
    "Cash": ["cash", "cash sales", "cash sale value"],
    "LDS": ["lds", "last day sales"],
    "PDR": ["pdr", "projected daily run rate", "required run rate"],
    "TDS": ["tds", "today sales", "today's sales"],
    "Branch Category": ["branch category", "performance band"],
    "Date": ["date", "transaction date", "sales date", "invoice date"],
    "SV": ["sv", "sale value", "sales value", "net sales"],
    "QTY": ["qty", "quantity", "units", "units sold"],
    "LINEAMOUNT": ["lineamount", "line amount", "amount"],
    "SALESPRICE": ["salesprice", "sales price", "unit price"]
}

def norm_name(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value).lower()).strip()

def auto_map(columns, fields):
    normalized = {norm_name(c): c for c in columns}
    mapping = {}
    for field in fields:
        candidates = [field] + CANONICAL_ALIASES.get(field, [])
        found = None
        for alias in candidates:
            if norm_name(alias) in normalized:
                found = normalized[norm_name(alias)]
                break
        mapping[field] = found
    return mapping

def apply_mapping(df, mapping):
    return df.rename(columns={src: dst for dst, src in mapping.items() if src})


def excel_like_filters(df, key_prefix):
    """Return a filtered dataframe using per-column controls similar to Excel."""
    filtered = df.copy()

    with st.expander("Column Filters", expanded=False):
        st.caption("Choose filters for any column. Leave a filter unchanged to show all rows.")

        for idx, col in enumerate(filtered.columns):
            series = filtered[col]
            control_key = f"{key_prefix}_{idx}_{col}"

            # Datetime columns
            if pd.api.types.is_datetime64_any_dtype(series):
                valid = series.dropna()
                if valid.empty:
                    continue
                min_date = valid.min().date()
                max_date = valid.max().date()
                selected = st.date_input(
                    f"{col}",
                    value=(min_date, max_date),
                    min_value=min_date,
                    max_value=max_date,
                    key=control_key
                )
                if isinstance(selected, tuple) and len(selected) == 2:
                    start, end = selected
                    filtered = filtered[
                        filtered[col].dt.date.between(start, end)
                    ]
                continue

            # Numeric columns
            numeric_series = pd.to_numeric(series, errors="coerce")
            numeric_ratio = numeric_series.notna().mean()
            if numeric_ratio >= 0.8 and numeric_series.notna().any():
                min_val = float(numeric_series.min())
                max_val = float(numeric_series.max())
                if min_val == max_val:
                    st.number_input(f"{col}", value=min_val, disabled=True, key=control_key)
                else:
                    selected_range = st.slider(
                        f"{col}",
                        min_value=min_val,
                        max_value=max_val,
                        value=(min_val, max_val),
                        key=control_key
                    )
                    filtered = filtered[
                        pd.to_numeric(filtered[col], errors="coerce").between(
                            selected_range[0], selected_range[1]
                        )
                    ]
                continue

            # Categorical columns with manageable unique values
            unique_vals = series.dropna().astype(str).unique().tolist()
            unique_vals = sorted(unique_vals)
            if 0 < len(unique_vals) <= 60:
                selected_vals = st.multiselect(
                    f"{col}",
                    options=unique_vals,
                    default=unique_vals,
                    key=control_key
                )
                if selected_vals:
                    filtered = filtered[filtered[col].astype(str).isin(selected_vals)]
                else:
                    filtered = filtered.iloc[0:0]
                continue

            # Free-text filter for high-cardinality columns
            text_query = st.text_input(
                f"{col} contains",
                key=control_key,
                placeholder="Type text to filter..."
            )
            if text_query:
                filtered = filtered[
                    filtered[col].astype(str).str.contains(
                        text_query, case=False, na=False
                    )
                ]

    return filtered


@st.cache_data
def load_workbook(source):
    working = pd.read_excel(source, sheet_name="Working Sheet", header=9)
    raw = pd.read_excel(source, sheet_name="Raw Data")
    working = working.loc[:, ~working.columns.astype(str).str.startswith("Unnamed")]
    return working, raw

def num(series):
    return pd.to_numeric(series, errors="coerce").fillna(0)

def money(v):
    v = float(v)
    if abs(v) >= 1_000_000_000:
        return f"Rs. {v/1_000_000_000:.2f}B"
    if abs(v) >= 1_000_000:
        return f"Rs. {v/1_000_000:.1f}M"
    if abs(v) >= 1_000:
        return f"Rs. {v/1_000:.1f}K"
    return f"Rs. {v:,.0f}"

def pct(v):
    return f"{float(v)*100:.1f}%"

def safe_ratio(a, b):
    return a / b if b else 0

def excel_safe(value):
    """Convert pandas/numpy values into values that openpyxl can safely write."""
    if hasattr(value, "item"):
        value = value.item()
    if pd.isna(value):
        return None
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    return value

def prepare_working(df):
    out = df.copy()
    needed = [
        "Unit","Store","Zone","Branch Name","Target","Sale Value","ACH%",
        "LDS","PDR","TDS","YTD TGT","YTD SV","YTD ACH",
        "CS TGT","HP TGT","Cash","HP","HP ACH","Branch Category"
    ]
    for c in needed:
        if c not in out.columns:
            out[c] = 0 if c not in ["Unit","Store","Zone","Branch Name","Branch Category"] else ""
    out = out[out["Store"].notna()].copy()
    for c in ["Target","Sale Value","ACH%","LDS","PDR","TDS","YTD TGT","YTD SV","YTD ACH",
              "CS TGT","HP TGT","Cash","HP","HP ACH"]:
        out[c] = num(out[c])
    return out

def prepare_raw(df):
    out = df.copy()
    for c in ["SV","QTY","LINEAMOUNT","SALESPRICE"]:
        if c in out.columns:
            out[c] = num(out[c])
    for c in ["Date","INVOICEDATE","SALESDATE"]:
        if c in out.columns:
            out[c] = pd.to_datetime(out[c], errors="coerce")
    return out

def make_excel_report(zone, top, metrics):
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    ws["A1"] = "Executive Sales Intelligence"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="223F72")
    r = 3
    for k, v in metrics.items():
        ws.cell(r,1,k).font = Font(bold=True)
        ws.cell(r,2,v)
        r += 1

    ws2 = wb.create_sheet("Zone Performance")
    for j, c in enumerate(zone.columns,1):
        ws2.cell(1,j,c).font = Font(bold=True, color="FFFFFF")
        ws2.cell(1,j).fill = PatternFill("solid", fgColor="223F72")
    for i, row in enumerate(zone.itertuples(index=False),2):
        for j, v in enumerate(row,1):
            ws2.cell(i,j,excel_safe(v))

    ws3 = wb.create_sheet("Top Contributors")
    for j, c in enumerate(top.columns,1):
        ws3.cell(1,j,c).font = Font(bold=True, color="FFFFFF")
        ws3.cell(1,j).fill = PatternFill("solid", fgColor="223F72")
    for i, row in enumerate(top.itertuples(index=False),2):
        for j, v in enumerate(row,1):
            ws3.cell(i,j,excel_safe(v))

    for sheet in wb.worksheets:
        for col in sheet.columns:
            width = min(max(len(str(c.value)) if c.value is not None else 0 for c in col)+2, 35)
            sheet.column_dimensions[col[0].column_letter].width = width

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def make_pdf(metrics, insights):
    bio = io.BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    w,h = A4
    y = h-55
    c.setFont("Helvetica-Bold", 16)
    c.drawString(45,y,"Executive Sales Intelligence")
    y -= 28
    c.setFont("Helvetica",9)
    for k,v in metrics.items():
        c.drawString(50,y,f"{k}: {v}")
        y -= 15
    y -= 10
    c.setFont("Helvetica-Bold",12)
    c.drawString(45,y,"Management Insights")
    y -= 20
    c.setFont("Helvetica",9)
    for insight in insights:
        line = "• " + insight
        while len(line) > 95:
            cut = line[:95].rfind(" ")
            c.drawString(50,y,line[:cut])
            line = line[cut+1:]
            y -= 14
        c.drawString(50,y,line)
        y -= 18
    c.save()
    return bio.getvalue()

st.markdown('<div class="topbar">Executive Sales Intelligence Dashboard</div>', unsafe_allow_html=True)
st.markdown('<div class="note"><b>Excel-style column filters:</b> open the filter panel above any data table to filter text, categories, dates, or numeric ranges.</div>', unsafe_allow_html=True)
st.markdown('<div class="note"><b>Flexible mapping:</b> the app can recognize common alternative column names and lets you map unfamiliar columns manually.</div>', unsafe_allow_html=True)
st.markdown('<div class="note"><b>Single-file workflow:</b> Upload only one Excel workbook. The app reads both the <code>Working Sheet</code> and <code>Raw Data</code> tabs from that same file.</div>', unsafe_allow_html=True)


with st.sidebar:
    st.header("Data")
    source_mode = st.radio("Source", ["Bundled demo workbook", "Upload one workbook"])
    if source_mode == "Bundled demo workbook":
        source = DATA_FILE
        st.success("Workbook loaded")
    else:
        uploaded = st.file_uploader("Upload a single Excel workbook containing the required sheets", type=["xlsx"])
        if uploaded is None:
            st.stop()
        source = uploaded

working, raw = load_workbook(source)

with st.sidebar:
    st.divider()
    enable_mapping = st.checkbox("Enable flexible column mapping", value=True)

if enable_mapping:
    working_fields = ["Unit","Store","Zone","Branch Name","Target","Sale Value","LDS","PDR","TDS","YTD TGT","YTD SV","CS TGT","Cash","Branch Category"]
    raw_fields = ["Date","SV","QTY","LINEAMOUNT","SALESPRICE"]

    suggested_working = auto_map(working.columns.tolist(), working_fields)
    suggested_raw = auto_map(raw.columns.tolist(), raw_fields)

    with st.expander("Map workbook columns", expanded=False):
        st.caption("Suggested matches are selected automatically. Change any incorrect mapping.")
        working_mapping = {}
        for field in working_fields:
            options = ["— Not mapped —"] + working.columns.astype(str).tolist()
            suggestion = suggested_working.get(field)
            default = options.index(str(suggestion)) if suggestion in working.columns else 0
            choice = st.selectbox(f"Working Sheet → {field}", options, index=default, key=f"w_{field}")
            working_mapping[field] = None if choice == "— Not mapped —" else choice

        raw_mapping = {}
        for field in raw_fields:
            options = ["— Not mapped —"] + raw.columns.astype(str).tolist()
            suggestion = suggested_raw.get(field)
            default = options.index(str(suggestion)) if suggestion in raw.columns else 0
            choice = st.selectbox(f"Raw Data → {field}", options, index=default, key=f"r_{field}")
            raw_mapping[field] = None if choice == "— Not mapped —" else choice

    working = apply_mapping(working, working_mapping)
    raw = apply_mapping(raw, raw_mapping)

working = prepare_working(working)
raw = prepare_raw(raw)

units = sorted([x for x in working["Unit"].dropna().astype(str).unique() if x and x != "nan"])
zones = sorted([x for x in working["Zone"].dropna().astype(str).unique() if x and x != "nan"])

with st.sidebar:
    selected_units = st.multiselect("Unit", units, default=units)
    selected_zones = st.multiselect("Zone", zones, default=zones)

filtered = working[
    working["Unit"].astype(str).isin(selected_units) &
    working["Zone"].astype(str).isin(selected_zones)
].copy()

if filtered.empty:
    st.warning("No records match the selected filters.")
    st.stop()

# Date logic
date_candidates = []
for col in ["Date","INVOICEDATE","SALESDATE"]:
    if col in raw.columns:
        date_candidates.extend(raw[col].dropna().tolist())
as_of = max(date_candidates) if date_candidates else pd.Timestamp.today()
days_in_month = monthrange(as_of.year, as_of.month)[1]
elapsed_days = max(as_of.day, 1)
remaining_days = max(days_in_month - elapsed_days, 0)

# Core KPIs
mtd_target = filtered["Target"].sum()
mtd_sales = filtered["Sale Value"].sum()
mtd_ach = safe_ratio(mtd_sales, mtd_target)

cash_target = filtered["CS TGT"].sum()
cash_sales = filtered["Cash"].sum()
cash_ach = safe_ratio(cash_sales, cash_target)

ytd_target = filtered["YTD TGT"].sum()
ytd_sales = filtered["YTD SV"].sum()
ytd_ach = safe_ratio(ytd_sales, ytd_target)

required_run_rate = safe_ratio(max(mtd_target-mtd_sales,0), remaining_days)
current_run_rate = safe_ratio(mtd_sales, elapsed_days)
run_ratio = safe_ratio(current_run_rate, required_run_rate)
risk = "Low" if run_ratio >= 1 else ("Medium" if run_ratio >= .8 else "High")

# Summary cards
a,b,c,d = st.columns(4)
a.metric("MTD Sales", money(mtd_sales), f"Target {money(mtd_target)}")
b.metric("Cash Sales", money(cash_sales), f"Target {money(cash_target)}")
c.metric("YTD Sales", money(ytd_sales), f"Achievement {ytd_ach*100:.1f}%")
d.metric("Days Remaining", remaining_days, f"Risk: {risk}")

a2,b2,c2,d2 = st.columns(4)
a2.metric("MTD Achievement", f"{mtd_ach*100:.1f}%")
b2.metric("Cash Achievement", f"{cash_ach*100:.1f}%")
c2.metric("Required Run Rate", money(required_run_rate))
d2.metric("Current Run Rate", money(current_run_rate))

st.markdown('<div class="section-title">Zone Performance</div>', unsafe_allow_html=True)

zone = filtered.groupby("Zone", as_index=False).agg(
    **{
        "MTD TGT":("Target","sum"),
        "MTD SV":("Sale Value","sum"),
        "CS TGT":("CS TGT","sum"),
        "CS SV":("Cash","sum"),
        "LDS":("LDS","sum"),
        "PDR":("PDR","sum"),
        "TDS":("TDS","sum"),
        "YTD TGT":("YTD TGT","sum"),
        "YTD SV":("YTD SV","sum"),
    }
)
zone["MTD ACH"] = (zone["MTD SV"] / zone["MTD TGT"].replace(0,pd.NA)).fillna(0)
zone["CS ACH"] = (zone["CS SV"] / zone["CS TGT"].replace(0,pd.NA)).fillna(0)
zone["YTD ACH"] = (zone["YTD SV"] / zone["YTD TGT"].replace(0,pd.NA)).fillna(0)

zone_filtered = excel_like_filters(zone, "zone_table")
zone_display = zone_filtered.copy()
for col in ["MTD TGT","MTD SV","CS TGT","CS SV","LDS","PDR","TDS","YTD TGT","YTD SV"]:
    if col in zone_display.columns:
        zone_display[col] = zone_display[col].map(money)
for col in ["MTD ACH","CS ACH","YTD ACH"]:
    if col in zone_display.columns:
        zone_display[col] = zone_display[col].map(lambda x: f"{x*100:.1f}%")
st.dataframe(zone_display, use_container_width=True, hide_index=True)

left,right = st.columns([1.35,1])

with left:
    
st.markdown('<div class="section-title">Branch-Level Detail</div>', unsafe_allow_html=True)
branch_detail_columns = [
    "Unit","Store","Zone","Branch Name","Target","Sale Value","ACH%",
    "CS TGT","Cash","HP","LDS","PDR","TDS","YTD TGT","YTD SV","YTD ACH",
    "Branch Category"
]
branch_detail = filtered[[c for c in branch_detail_columns if c in filtered.columns]].copy()
branch_detail_filtered = excel_like_filters(branch_detail, "branch_detail")
st.dataframe(branch_detail_filtered, use_container_width=True, hide_index=True)

st.markdown('<div class="section-title">Road to Target</div>', unsafe_allow_html=True)
    road = zone[["Zone","MTD TGT","MTD SV"]].melt(
        id_vars="Zone",
        value_vars=["MTD TGT","MTD SV"],
        var_name="Measure",
        value_name="Value"
    )
    fig = px.bar(road, x="Zone", y="Value", color="Measure", barmode="group")
    fig.update_layout(yaxis_title="PKR", legend_title="")
    st.plotly_chart(fig, use_container_width=True)

with right:
    st.markdown('<div class="section-title">Branch Categorization</div>', unsafe_allow_html=True)
    cats = ["Below 15%","15%-35%","36%-55%","56%-85%","86%-95%","Above 95%"]
    cat_table = pd.crosstab(filtered["Zone"], filtered["Branch Category"])
    cat_table = cat_table.reindex(columns=cats, fill_value=0)
    st.dataframe(cat_table, use_container_width=True)

left2,right2 = st.columns([1.35,1])

with left2:
    st.markdown('<div class="section-title">YTD & MTD Achievement</div>', unsafe_allow_html=True)
    ach = zone[["Zone","MTD ACH","YTD ACH"]].melt(
        id_vars="Zone",
        value_vars=["MTD ACH","YTD ACH"],
        var_name="Metric",
        value_name="Achievement"
    )
    ach["Achievement"] *= 100
    fig = px.line(ach, x="Zone", y="Achievement", color="Metric", markers=True)
    fig.update_layout(yaxis_title="Achievement %")
    st.plotly_chart(fig, use_container_width=True)

with right2:
    st.markdown('<div class="section-title">Highest Value Contributors</div>', unsafe_allow_html=True)
    top = (
        filtered.groupby("Branch Name", as_index=False)["Sale Value"]
        .sum()
        .sort_values("Sale Value", ascending=False)
        .head(5)
    )
    top_filtered = excel_like_filters(top, "top_contributors")
    top_display = top_filtered.copy()
    top_display["Sale Value"] = top_display["Sale Value"].map(money)
    st.dataframe(top_display, use_container_width=True, hide_index=True)

st.markdown('<div class="section-title">Daily Tracking</div>', unsafe_allow_html=True)

date_col = "Date" if "Date" in raw.columns else ("INVOICEDATE" if "INVOICEDATE" in raw.columns else "SALESDATE")
value_col = "SV" if "SV" in raw.columns else ("SALESPRICE" if "SALESPRICE" in raw.columns else "LINEAMOUNT")
if date_col in raw.columns and value_col in raw.columns:
    raw_month = raw[(raw[date_col].dt.year == as_of.year) & (raw[date_col].dt.month == as_of.month)].copy()
    daily = raw_month.groupby(raw_month[date_col].dt.day)[value_col].sum()
    tracker = pd.DataFrame({"Day": range(1,days_in_month+1)})
    tracker["Daily Sales"] = tracker["Day"].map(daily).fillna(0)
    tracker["Cumulative Sales"] = tracker["Daily Sales"].cumsum()
    tracker["Achievement %"] = tracker["Cumulative Sales"] / mtd_target * 100 if mtd_target else 0
    tracker["Expected %"] = tracker["Day"] / days_in_month * 100

    fig = px.line(
        tracker,
        x="Day",
        y=["Achievement %","Expected %"],
        markers=True
    )
    fig.update_layout(yaxis_title="Percent")
    st.plotly_chart(fig, use_container_width=True)
    tracker_filtered = excel_like_filters(tracker, "daily_tracker")
    st.dataframe(
        tracker_filtered.style.format({
            "Daily Sales":"{:,.0f}",
            "Cumulative Sales":"{:,.0f}",
            "Achievement %":"{:.1f}%",
            "Expected %":"{:.1f}%"
        }),
        use_container_width=True,
        hide_index=True
    )

st.markdown('<div class="section-title">Management Insights</div>', unsafe_allow_html=True)
best_zone = zone.sort_values("MTD ACH", ascending=False).iloc[0]
worst_zone = zone.sort_values("MTD ACH").iloc[0]
insights = [
    f"{best_zone['Zone']} is the strongest zone on MTD achievement at {best_zone['MTD ACH']*100:.1f}%.",
    f"{worst_zone['Zone']} has the lowest MTD achievement at {worst_zone['MTD ACH']*100:.1f}%.",
    f"Current daily run rate is {money(current_run_rate)} versus a required run rate of {money(required_run_rate)}.",
    f"Cash achievement is {cash_ach*100:.1f}% and YTD achievement is {ytd_ach*100:.1f}%.",
    f"The current projected operational risk level is {risk}.",
]
for insight in insights:
    st.info(insight)

metrics = {
    "As of Date": as_of.strftime("%d-%b-%Y"),
    "MTD Sales": money(mtd_sales),
    "MTD Target": money(mtd_target),
    "MTD Achievement": f"{mtd_ach*100:.1f}%",
    "Cash Sales": money(cash_sales),
    "Cash Target": money(cash_target),
    "Cash Achievement": f"{cash_ach*100:.1f}%",
    "YTD Sales": money(ytd_sales),
    "YTD Target": money(ytd_target),
    "YTD Achievement": f"{ytd_ach*100:.1f}%",
    "Remaining Days": remaining_days,
    "Required Run Rate": money(required_run_rate),
    "Current Run Rate": money(current_run_rate),
    "Risk Level": risk,
}

st.markdown('<div class="section-title">Reports</div>', unsafe_allow_html=True)
x1,x2 = st.columns(2)
x1.download_button(
    "Download Excel Report",
    make_excel_report(zone, top, metrics),
    "executive_sales_report.xlsx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True
)
x2.download_button(
    "Download PDF Summary",
    make_pdf(metrics, insights),
    "executive_sales_summary.pdf",
    "application/pdf",
    use_container_width=True
)

st.caption("Automated dashboard generated from the uploaded workbook.")
